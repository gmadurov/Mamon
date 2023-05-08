from django.http import HttpResponse
from django.core.exceptions import FieldDoesNotExist
from django.utils import timezone

from openpyxl import Workbook
import csv

from purchase.models import Purchase


def purchase_export_excel(modeladmin, request, queryset, name=''):
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "".join(
        [
            "attachment; filename=",
            name or timezone.localtime().strftime("%Y%m%d_%H%M%S"),
            "_mamon_purchases_export.xlsx",
        ]
    )
    exportwb = Workbook()
    exportws = exportwb.active

    # Define fieldlist for purchasefields
    feilds = [
        "created",
        "seller",
        "orders",
        "total",
        "payment_method",
    ]

    # Define fieldlist for Lidmaatschapfields

    for j, f in enumerate(feilds):
        exportws.cell(row=1, column=j + 1).value = f.title()

    for i, purchase in enumerate(queryset):
        for j, f in enumerate(feilds):
            try:
                if Purchase._meta.get_field(f).many_to_many or Purchase._meta.get_field(f).one_to_many:
                    if f == "orders":
                        exportws.cell(row=i + 2, column=j + 1).value = ", ".join([q.name for q in getattr(purchase, "orders").all()])
                    else:
                        exportws.cell(row=i + 2, column=j + 1).value = ", ".join([str(q) for q in getattr(purchase, f).all()])
                elif Purchase._meta.get_field(f).choices:
                    exportws.cell(row=i + 2, column=j + 1).value = getattr(purchase, "get_" + f + "_display")()
                else:
                    try:
                        if f == "created":
                            exportws.cell(row=i + 2, column=j + 1).value = getattr(purchase, f).strftime("%Y-%m-%d %H:%M")
                        else:
                            exportws.cell(row=i + 2, column=j + 1).value = getattr(purchase, f)
                    except ValueError:
                        exportws.cell(row=i + 2, column=j + 1).value = str(getattr(purchase, f)).replace("None", "")
            except FieldDoesNotExist as e:
                if getattr(purchase, f).__class__.__name__ == "method":
                    exportws.cell(row=i + 2, column=j + 1).value = getattr(purchase, f)()
                elif f == "total":
                    exportws.cell(row=i + 2, column=j + 1).value = purchase.total
                elif f == "payment_method":
                    exportws.cell(row=i + 2, column=j + 1).value = purchase.payment_method
                    

    exportwb.save(response)
    return response


purchase_export_excel.short_description = "Exporteer Excel (XLSX)"
